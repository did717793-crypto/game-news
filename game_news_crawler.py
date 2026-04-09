#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
게임 업계 동향 크롤러 v2.1
- 10개 사이트 크롤링 (매일 00:00 KST 실행, 주말 포함)
- 수집 범위: 실행 전일 00:00 ~ 23:59 KST (고정 24h)
- XLSX 출력 (기존 구조 유지: 신작 소식 / 게임 회사 동향 / HOT TOPIC / 일반)
- HTML 출력 (신규: 신작 소식 / 게임 소식 / 게임 회사 동향 / 일반) → GitHub Pages 배포
- 실행 환경: Python 3.10+
"""

import os
import re
import json
import subprocess
import concurrent.futures
from datetime import datetime, timedelta
from pathlib import Path

import requests
import feedparser
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pytz

# ──────────────────────────────────────────────────────────────────────────────
# 설정
# ──────────────────────────────────────────────────────────────────────────────
KST = pytz.timezone("Asia/Seoul")

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)
HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}
# 기사 본문 fetch 전용 헤더 (브라우저 접속처럼 위장)
ARTICLE_HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Cache-Control": "max-age=0",
}

# 저장 경로
BASE_DIR   = Path(r"C:\Users\Admin\Desktop\일일 체크리스트")
GITHUB_DIR = Path(r"C:\Users\Admin\Documents\game-news")  # git clone 위치

# 루리웹 HOT TOPIC 조회수 기준
RULIWEB_HOT_VIEWS = 5000


# ──────────────────────────────────────────────────────────────────────────────
# 유틸리티
# ──────────────────────────────────────────────────────────────────────────────
def clean(text: str) -> str:
    """surrogate / 제어문자 제거 및 공백 정리"""
    if not isinstance(text, str):
        text = str(text) if text else ""
    text = text.encode("utf-8", errors="ignore").decode("utf-8", errors="ignore")
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
    return re.sub(r'\s+', ' ', text).strip()


def get_now() -> datetime:
    return datetime.now(KST)


def get_content_date(now: datetime) -> datetime:
    """수집 대상일 = 실행일 전날 (00:00 KST 기준)"""
    yesterday = now - timedelta(days=1)
    # 자정(00:00)으로 초기화
    return yesterday.replace(hour=0, minute=0, second=0, microsecond=0)


def get_time_window(now: datetime):
    """수집 윈도우: 전일 00:00:00 ~ 23:59:59 KST (항상 24h 고정)"""
    start = get_content_date(now)
    end   = start + timedelta(hours=23, minutes=59, seconds=59)
    return start, end


def get_folder_name(content_date: datetime) -> str:
    """폴더명은 수집 대상일(전일) 기준"""
    return content_date.strftime("%Y.%m.%d")


def parse_pub_dt(pub_dt_str) -> "datetime | None":
    """RSS published 문자열 -> KST datetime 변환"""
    if not pub_dt_str:
        return None
    try:
        import email.utils
        dt = email.utils.parsedate_to_datetime(pub_dt_str)
        return dt.astimezone(KST)
    except Exception:
        return None


def fetch(url: str, timeout: int = 15):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return r
    except Exception as e:
        print(f"  [WARN] fetch 실패: {url} — {e}")
        return None


def parse_rss(url: str) -> list:
    try:
        feed = feedparser.parse(url)
        return feed.entries
    except Exception as e:
        print(f"  [WARN] RSS 파싱 실패: {url} — {e}")
        return []


# ──────────────────────────────────────────────────────────────────────────────
# 카테고리 분류 키워드
# ──────────────────────────────────────────────────────────────────────────────
KW_NEW_GAME = [
    "신작", "출시", "런칭", "오픈", "사전예약", "베타", "cbt", "obt",
    "공개", "골드행", "발매", "서비스 시작", "서비스 오픈", "얼리액세스",
    "launch", "release", "announce", "reveal", "pre-registration",
    "early access", "open beta", "closed beta",
]

KW_COMPANY = [
    "인수", "합병", "투자", "펀딩", "실적", "매출", "영업이익", "순이익",
    "구조조정", "서비스 종료", "서비스종료", "파트너십", "계약", "지분",
    "상장", "ipo", "직원 감축", "정리해고",
    "acquisition", "layoff", "merger", "funding", "raises", "acquires",
    "bankruptcy", "shutdown", "restructur", "revenue", "earnings",
]

KW_GAME_NEWS = [
    "업데이트", "패치", "밸런스", "신규 서버", "신규서버", "이벤트",
    "콘텐츠", "시즌", "리뉴얼", "점검", "대규모", "새로운 챕터",
    "신규 캐릭터", "신규 영웅", "신규 맵", "신규 던전",
    "update", "patch", "balance", "season", "content", "event",
    "new character", "new map", "expansion", "dlc", "hotfix",
]


def classify_html(title: str, summary: str) -> str:
    """HTML 출력용 4개 카테고리 분류"""
    text = (title + " " + summary).lower()
    for kw in KW_NEW_GAME:
        if kw in text:
            return "신작 소식"
    for kw in KW_COMPANY:
        if kw in text:
            return "게임 회사 동향"
    for kw in KW_GAME_NEWS:
        if kw in text:
            return "게임 소식"
    return "일반"


def classify_xlsx(title: str, summary: str, is_ruliweb_best: bool, views: int) -> str:
    """XLSX 출력용 원본 4개 카테고리 분류"""
    text = (title + " " + summary).lower()
    for kw in KW_NEW_GAME:
        if kw in text:
            return "신작 소식"
    for kw in KW_COMPANY:
        if kw in text:
            return "게임 회사 동향"
    if is_ruliweb_best or (views and views >= RULIWEB_HOT_VIEWS):
        return "HOT TOPIC"
    return "일반"


# ──────────────────────────────────────────────────────────────────────────────
# 장르 감지 + 제목 접두어 제거
# ──────────────────────────────────────────────────────────────────────────────
_GENRE_KW = [
    ("MMORPG",     ["mmorpg", "mmo", "다중접속", "대규모 다중"]),
    ("수집형 RPG", ["수집형", "가챠", "컬렉션 rpg"]),
    ("방치형",     ["방치형", "방치 rpg", "idle"]),
    ("전략",       ["전략", "slg", "전략 시뮬"]),
    ("슈팅",       ["슈팅", "fps", "tps"]),
    ("배틀로얄",   ["배틀로얄", "battle royale"]),
    ("스포츠",     ["스포츠", "풋볼", "축구"]),
    ("액션 RPG",   ["액션 rpg", "action rpg"]),
    ("RPG",        ["rpg"]),
]

def detect_genre(title: str, body: str) -> str:
    text = (title + " " + (body or "")[:300]).lower()
    for genre, kws in _GENRE_KW:
        if any(kw in text for kw in kws):
            return genre
    return ""

_PREFIX_RE = re.compile(r'^(\[.*?\]|\(.*?\)|【.*?】|《.*?》|〈.*?〉)\s*')

def strip_title_prefix(title: str) -> str:
    """[단독], [게임동향] 등 제목 앞 접두어 반복 제거"""
    result = title
    while True:
        m = _PREFIX_RE.match(result)
        if m:
            result = result[m.end():]
        else:
            break
    return result.strip()


# ──────────────────────────────────────────────────────────────────────────────
# 기사 저장소
# ──────────────────────────────────────────────────────────────────────────────
ARTICLES: list[dict] = []
_seen_urls:   set[str] = set()
_seen_titles: set[str] = set()


def add_article(article: dict):
    url = article.get("url", "").strip()
    title = article.get("title", "").strip()

    if not url or not title:
        return
    if url in _seen_urls:
        return

    title_key = re.sub(r'\s', '', title[:30]).lower()
    if title_key in _seen_titles:
        return

    _seen_urls.add(url)
    _seen_titles.add(title_key)

    article["title"]        = clean(title)
    article["summary"]      = clean(article.get("summary", ""))
    article["site"]         = clean(article.get("site", ""))
    article["collected_at"] = get_now().strftime("%Y-%m-%d %H:%M")
    article.setdefault("views", 0)
    article.setdefault("comments", 0)
    article.setdefault("pub_dt", None)
    article.setdefault("is_domestic", True)
    article.setdefault("is_ruliweb_best", False)

    # pub_dt 파싱 → pub_date (표시용) + _pub_datetime (날짜 필터용)
    _pdt = parse_pub_dt(article.get("pub_dt"))
    article["_pub_datetime"] = _pdt
    article["pub_date"] = _pdt.strftime("%Y-%m-%d %H:%M") if _pdt else ""

    ARTICLES.append(article)


# ──────────────────────────────────────────────────────────────────────────────
# 크롤러 (10개 사이트)
# ──────────────────────────────────────────────────────────────────────────────

# 1. 디스이즈게임 — Google News RSS (Cloudflare 차단으로 직접 접근 불가)
def crawl_thisisgame():
    print("  ▶ 디스이즈게임")
    url = "https://news.google.com/rss/search?q=site:thisisgame.com&hl=ko&gl=KR&ceid=KR:ko"
    for e in parse_rss(url)[:25]:
        add_article({
            "site": "디스이즈게임",
            "title": e.get("title", ""),
            "url": e.get("link", ""),
            "summary": clean(e.get("summary", ""))[:200],
            "pub_dt": e.get("published", None),
            "is_domestic": True,
        })


# 2. 게임메카 — HTML 파싱
def crawl_gamemeca():
    print("  ▶ 게임메카")
    r = fetch("https://www.gamemeca.com/news.php")
    if not r:
        return
    soup = BeautifulSoup(r.text, "lxml")
    for a in soup.select('a[href*="view.php?gid="]')[:25]:
        title = clean(a.get_text())
        if not title or len(title) < 5:
            continue
        href = a.get("href", "")
        if not href.startswith("http"):
            href = "https://www.gamemeca.com/" + href.lstrip("/")
        add_article({"site": "게임메카", "title": title, "url": href, "is_domestic": True})


# 3. 게임조선 — HTML 파싱
def crawl_gamechosun():
    print("  ▶ 게임조선")
    r = fetch("https://gamechosun.co.kr/")
    if not r:
        return
    soup = BeautifulSoup(r.text, "lxml")
    anchors = (
        soup.select('a[href*="view.php?no="]') +
        soup.select('a[href*="/article/"]')
    )
    seen_in_page: set[str] = set()
    for a in anchors[:30]:
        title = clean(a.get_text())
        if not title or len(title) < 5:
            continue
        href = a.get("href", "")
        if not href.startswith("http"):
            href = "https://gamechosun.co.kr" + href
        if href in seen_in_page:
            continue
        seen_in_page.add(href)
        add_article({"site": "게임조선", "title": title, "url": href, "is_domestic": True})


# 4. 인벤 — RSS
def crawl_inven():
    print("  ▶ 인벤")
    for e in parse_rss("https://www.inven.co.kr/webzine/news/rss.php")[:25]:
        add_article({
            "site": "인벤",
            "title": e.get("title", ""),
            "url": e.get("link", ""),
            "summary": clean(e.get("summary", ""))[:200],
            "pub_dt": e.get("published", None),
            "is_domestic": True,
        })


# 5. 게임포커스 — HTML 파싱
def crawl_gamefocus():
    print("  ▶ 게임포커스")
    r = fetch("https://www.gamefocus.co.kr/")
    if not r:
        return
    soup = BeautifulSoup(r.text, "lxml")
    for a in soup.select('a[href*="detail.php?number="]')[:25]:
        title = clean(a.get_text())
        if not title or len(title) < 5:
            continue
        href = a.get("href", "")
        if not href.startswith("http"):
            href = "https://www.gamefocus.co.kr/" + href.lstrip("/")
        add_article({"site": "게임포커스", "title": title, "url": href, "is_domestic": True})


# 6. 루리웹 — 뉴스 + 베스트 HTML 파싱
def crawl_ruliweb():
    print("  ▶ 루리웹 뉴스")
    r = fetch("https://bbs.ruliweb.com/news")
    if r:
        soup = BeautifulSoup(r.text, "lxml")
        for row in soup.select("table.board_list_table tr"):
            a = row.select_one("td.subject a.deco")
            if not a:
                continue
            title = clean(a.get_text())
            url = a.get("href", "")
            if not url.startswith("http"):
                url = "https://bbs.ruliweb.com" + url

            views = 0
            comments = 0
            vt = row.select_one("td.hit")
            if vt:
                try:
                    views = int(re.sub(r'[^\d]', '', vt.get_text()))
                except Exception:
                    pass
            ct = row.select_one("td.replycount")
            if ct:
                try:
                    comments = int(re.sub(r'[^\d]', '', ct.get_text()))
                except Exception:
                    pass

            add_article({
                "site": "루리웹",
                "title": title,
                "url": url,
                "is_domestic": True,
                "is_ruliweb_best": False,
                "views": views,
                "comments": comments,
            })

    print("  ▶ 루리웹 베스트")
    r2 = fetch("https://bbs.ruliweb.com/best/article/all")
    if r2:
        soup = BeautifulSoup(r2.text, "lxml")
        for row in soup.select("table.board_list_table tr"):
            a = row.select_one("td.subject a.deco")
            if not a:
                continue
            title = clean(a.get_text())
            url = a.get("href", "")
            if not url.startswith("http"):
                url = "https://bbs.ruliweb.com" + url

            views = 0
            vt = row.select_one("td.hit")
            if vt:
                try:
                    views = int(re.sub(r'[^\d]', '', vt.get_text()))
                except Exception:
                    pass

            add_article({
                "site": "루리웹",
                "title": title,
                "url": url,
                "is_domestic": True,
                "is_ruliweb_best": True,
                "views": views,
                "comments": 0,
            })


# 7. 네이버뉴스 IT/게임 섹션 — HTML 파싱
def crawl_naver():
    print("  ▶ 네이버뉴스 IT/게임")
    r = fetch("https://news.naver.com/section/105")
    if not r:
        return
    soup = BeautifulSoup(r.text, "lxml")
    for a in soup.select("a.sa_text_title")[:25]:
        title = clean(a.get_text())
        if not title or len(title) < 5:
            continue
        href = a.get("href", "")
        if not href.startswith("http"):
            href = "https://news.naver.com" + href
        add_article({"site": "네이버뉴스", "title": title, "url": href, "is_domestic": True})


# 8. Game Developer — RSS
def crawl_gamedeveloper():
    print("  ▶ Game Developer")
    for e in parse_rss("https://www.gamedeveloper.com/rss.xml")[:25]:
        add_article({
            "site": "Game Developer",
            "title": e.get("title", ""),
            "url": e.get("link", ""),
            "summary": clean(e.get("summary", ""))[:200],
            "pub_dt": e.get("published", None),
            "is_domestic": False,
        })


# 9. VentureBeat — RSS + 게임 키워드 필터
VB_GAME_KW = [
    "game", "gaming", "esport", "nintendo", "sony", "microsoft", "xbox",
    "playstation", "steam", "mobile game", "rpg", "mmorpg", "fps", "unity",
    "unreal", "epic games", "riot", "blizzard", "activision", "ubisoft",
    "ea games", "take-two", "netmarble", "nexon", "krafton", "ncsoft",
]


def crawl_venturebeat():
    print("  ▶ VentureBeat")
    for e in parse_rss("https://venturebeat.com/feed/")[:60]:
        text = (e.get("title", "") + " " + e.get("summary", "")).lower()
        if not any(kw in text for kw in VB_GAME_KW):
            continue
        add_article({
            "site": "VentureBeat",
            "title": e.get("title", ""),
            "url": e.get("link", ""),
            "summary": clean(e.get("summary", ""))[:200],
            "pub_dt": e.get("published", None),
            "is_domestic": False,
        })


# 10. PocketGamer.biz — HTML 파싱 (모바일 B2B)
def crawl_pocketgamer():
    print("  ▶ PocketGamer.biz")
    r = fetch("https://www.pocketgamer.biz/news/")
    if not r:
        return
    soup = BeautifulSoup(r.text, "lxml")
    for a in soup.select("article a")[:30]:
        title = clean(a.get_text())
        if not title or len(title) < 15:
            continue
        href = a.get("href", "")
        if not href.startswith("http"):
            href = "https://www.pocketgamer.biz" + href
        add_article({"site": "PocketGamer.biz", "title": title, "url": href, "is_domestic": False})


def run_all(win_start=None, win_end=None):
    print("\n[1/3] 크롤링 시작")
    crawl_thisisgame()
    crawl_gamemeca()
    crawl_gamechosun()
    crawl_inven()
    crawl_gamefocus()
    crawl_ruliweb()
    crawl_naver()
    crawl_gamedeveloper()
    crawl_venturebeat()
    crawl_pocketgamer()

    # ── 날짜 필터: pub_date 있는 기사는 수집 윈도우 내 것만 유지 ──────────────
    if win_start and win_end:
        ARTICLES[:] = [
            a for a in ARTICLES
            if a.get("_pub_datetime") is None  # 날짜 모름(HTML 스크래핑) → 포함
            or (win_start <= a["_pub_datetime"] <= win_end)
        ]

    # 카테고리 분류
    for art in ARTICLES:
        art["cat_html"] = classify_html(art["title"], art["summary"])
        art["cat_xlsx"] = classify_xlsx(
            art["title"],
            art["summary"],
            art.get("is_ruliweb_best", False),
            art.get("views", 0),
        )

    # ── 해외 기사 제거 (국내 언론사 기사만 대시보드에 표시) ──────────────────
    ARTICLES[:] = [a for a in ARTICLES if a.get("is_domestic", True)]

    # ── 루리웹 BEST 비게임 기사 필터링 ──────────────────────────────────────
    GAME_FILTER_KW = [
        "게임", "모바일", "스팀", "콘솔", "플스", "엑박", "ps5", "ps4", "xbox",
        "닌텐도", "서버", "패치", "업데이트", "출시", "런칭", "사전예약", "오픈",
        "던파", "메이플", "로스트아크", "로아", "리니지", "오버워치", "롤",
        "발로란트", "넥슨", "엔씨", "크래프톤", "카카오게임즈", "넷마블",
        "펄어비스", "스마일게이트", "컴투스", "위메이드", "시프트업",
        "배그", "배틀그라운드", "디아블로", "포트나이트", "마인크래프트",
        "mmo", "rpg", "fps", "moba", "pc방", "e스포츠", "esports",
    ]
    ARTICLES[:] = [
        a for a in ARTICLES
        if not a.get("is_ruliweb_best")
        or any(kw in a.get("title", "").lower() for kw in GAME_FILTER_KW)
    ]

    # ── 사업 PM 관점 콘텐츠 중요도 스코어링 ────────────────────────────────
    # Tier S+ (+12): 신작 발표·사전예약·서비스 오픈 — 경쟁사 게임 라이프사이클 핵심
    PM_TIER_SP = [
        "사전예약", "사전 등록", "사전예약 시작", "사전예약 오픈",
        "cbt", "obt", "얼리 액세스", "얼리엑세스",
        "그랜드 오픈", "신규 서버", "서비스 시작", "정식 출시", "정식출시",
        "정식 서비스", "오픈일 확정", "출시일 확정", "최초 공개", "신작 공개",
        "첫 공개", "월드 프리미어",
    ]
    # Tier S (+10): 업계 구조 변화 — 반드시 알아야 할 비즈니스 이벤트
    PM_TIER_S = [
        "인수", "합병", "매각", "파산", "폐업", "청산",
        "해고", "감원", "구조조정", "희망퇴직", "인력 감축", "정리해고",
        "대규모 해고", "직원 해고", "직원 감축",
        "상장", "ipo", "코스닥 상장", "코스피 상장",
        "서비스 종료", "서버 종료", "서비스종료", "게임 종료",
    ]
    # Tier A (+8): 재무·경영 이벤트
    PM_TIER_A = [
        "투자 유치", "투자", "시리즈 a", "시리즈 b", "시리즈 c",
        "영업이익", "영업손실", "영업 이익", "영업 손실",
        "매출", "적자", "흑자", "손익", "영업실적", "실적 발표",
        "대표이사 교체", "ceo 교체", "신규 대표", "대표 취임", "대표 사임",
    ]
    # Tier B (+6): 게임 업계 흐름·기술 트렌드 (AI는 아래 별도 처리)
    PM_TIER_B = [
        "언리얼", "유니티", "클라우드 게임",
        "게임 시장", "모바일 시장", "글로벌 진출", "해외 진출",
        "규제", "심의", "게임 법", "확률형 아이템",
        "e스포츠", "리그", "대회", "월드챔피언십",
    ]
    # AI는 게임 컨텍스트와 함께 있을 때만 Tier B 인정
    _AI_TERMS   = ["ai", "인공지능", "생성형", "llm", "머신러닝"]
    _AI_GAME_CTX= ["게임", "개발", "스튜디오", "게임사", "콘텐츠", "npc", "캐릭터"]
    # Tier C (+4): 출시·런칭 (일반)
    PM_TIER_C = [
        "출시", "런칭", "오픈", "공개", "발표",
    ]
    # Tier D (+2): 일반 게임 소식
    PM_TIER_D = [
        "업데이트", "패치", "신규 콘텐츠", "이벤트", "시즌",
    ]
    # Tier E (0): 게임 외적 상품 — site_cnt만 반영
    PM_TIER_E = [
        "피규어", "굿즈", "인형", "쿠션", "키링", "포스터", "아크릴",
        "마우스", "키보드", "헤드셋", "모니터", "메모리", "ram", "그래픽카드",
        "노트북", "데스크탑", "주변기기", "pc 부품", "하드웨어",
        "게이밍 체어", "게이밍 의자",
    ]
    # Tier PROMO (+2): 할인/무료 프로모션 — S+ 키워드와 겹쳐도 강제 하향
    PM_TIER_PROMO = [
        "무료 체험", "무료 플레이", "무료 배포", "무료로 즐길", "무료 증정",
        "할인 프로모션", "할인 이벤트", "할인 판매", "특별 할인",
        "무료 증정", "공짜로", "번들 증정",
    ]

    for art in ARTICLES:
        t = art.get("title", "").lower()
        if any(kw in t for kw in PM_TIER_E):
            cs = 0
        elif any(kw in t for kw in PM_TIER_PROMO):
            cs = 2          # 프로모션은 S+ 발동 전에 잡아서 Tier D 수준으로 고정
        elif any(kw in t for kw in PM_TIER_SP):
            cs = 12
        elif any(kw in t for kw in PM_TIER_S):
            cs = 10
        elif any(kw in t for kw in PM_TIER_A):
            cs = 8
        elif (any(kw in t for kw in _AI_TERMS)
              and any(kw in t for kw in _AI_GAME_CTX)):
            cs = 6          # AI + 게임 컨텍스트 → Tier B
        elif any(kw in t for kw in PM_TIER_B):
            cs = 6
        elif any(kw in t for kw in PM_TIER_C):
            cs = 4
        elif any(kw in t for kw in PM_TIER_D):
            cs = 2
        else:
            cs = 1
        art["_content_score"] = cs

    # ── 중복 기사 클러스터링 & 대표 1건 선택 ─────────────────────────────────
    # 3자 이상 고유 키워드 2개 이상 겹치면 동일 기사로 판단
    _STOP = {
        "게임", "이번", "이후", "공개", "발표", "서비스", "업데이트",
        "이상", "국내", "지난", "최근", "현재", "시작", "진행", "관련",
        "뉴스", "기사", "한국", "새로운", "출시", "오픈",
    }

    def _dedup_kw(title: str) -> set:
        raw = set(re.findall(r"[가-힣a-zA-Z0-9]{2,}", title.lower()))
        return {w for w in raw if len(w) >= 3 and w not in _STOP}

    kw_list = [_dedup_kw(a.get("title", "")) for a in ARTICLES]
    visited = [False] * len(ARTICLES)
    clusters = []
    for i in range(len(ARTICLES)):
        if visited[i]:
            continue
        cluster = [i]
        visited[i] = True
        for j in range(i + 1, len(ARTICLES)):
            if visited[j]:
                continue
            if len(kw_list[i] & kw_list[j]) >= 2:   # 고유 키워드 2개 이상 일치
                cluster.append(j)
                visited[j] = True
        clusters.append(cluster)

    kept = []
    for cluster in clusters:
        # 클러스터에서 content_score 가장 높은 기사를 대표로 선택
        # (같으면 수집 사이트 다양성 기준으로 우선)
        best = max(cluster, key=lambda i: (
            ARTICLES[i]["_content_score"],
            len(ARTICLES[i].get("site", "")),
        ))
        rep = ARTICLES[best]
        # 클러스터 사이트 목록 (대표 기사 + 중복들)
        covered = list({ARTICLES[idx].get("site", "") for idx in cluster
                        if ARTICLES[idx].get("site")})
        rep["_site_cnt"]      = len(covered)
        rep["_covered_sites"] = covered
        rep["_score"]         = rep["_site_cnt"] + rep["_content_score"]
        kept.append(rep)

    ARTICLES[:] = kept

    print(f"\n  중복 제거 후: {len(ARTICLES)}건 (원본 클러스터 {len(clusters)}개)")
    by_cat = {}
    for art in ARTICLES:
        c = art["cat_html"]
        by_cat[c] = by_cat.get(c, 0) + 1
    for k, v in by_cat.items():
        print(f"  - {k}: {v}건")


# ──────────────────────────────────────────────────────────────────────────────
# XLSX 출력 (기존 구조 유지)
# ──────────────────────────────────────────────────────────────────────────────
CAT_COLORS_XLSX = {
    "신작 소식":     "FFF2CC",
    "게임 회사 동향": "DDEEFF",
    "HOT TOPIC":    "FFE0E0",
    "일반":          "F8F8F8",
}


def write_ws(ws, articles: list[dict]):
    headers = ["카테고리", "출처", "제목", "요약", "URL", "조회수", "댓글", "수집시각"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1C2340")
        cell.alignment = Alignment(horizontal="center")

    for art in articles:
        cat = art.get("cat_xlsx", "일반")
        ws.append([
            cat,
            art.get("site", ""),
            art.get("title", ""),
            art.get("summary", ""),
            art.get("url", ""),
            art.get("views", 0),
            art.get("comments", 0),
            art.get("collected_at", ""),
        ])
        color = CAT_COLORS_XLSX.get(cat, "F8F8F8")
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill("solid", fgColor=color)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 52
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 16


def save_xlsx(content_date: datetime) -> Path:
    """파일·폴더명 모두 수집 대상일(전일) 기준"""
    folder_name = get_folder_name(content_date)
    folder_path = BASE_DIR / folder_name
    folder_path.mkdir(parents=True, exist_ok=True)

    file_path = folder_path / f"game_news_{content_date.strftime('%Y%m%d')}.xlsx"

    wb = openpyxl.Workbook()

    ws_all = wb.active
    ws_all.title = "전체"
    write_ws(ws_all, ARTICLES)

    for cat_name in ["신작 소식", "게임 회사 동향", "HOT TOPIC"]:
        ws = wb.create_sheet(title=cat_name)
        write_ws(ws, [a for a in ARTICLES if a.get("cat_xlsx") == cat_name])

    wb.save(file_path)
    print(f"  XLSX 저장: {file_path}")
    return file_path


# ──────────────────────────────────────────────────────────────────────────────
# HTML 출력 (lol.ps 스타일, 신규)
# ──────────────────────────────────────────────────────────────────────────────
HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="ko" data-theme="light">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>GAME PULSE — PM 대시보드</title>
  <style>
    *,*::before,*::after{box-sizing:border-box;margin:0;padding:0;}

    /* ── CSS VARIABLES (light default) ── */
    :root {
      --bg:#f4f5f7;
      --surface:#ffffff;
      --surface2:#f0f1f4;
      --border:#dde1e7;
      --text:#1e1e2e;
      --text2:#555770;
      --text3:#8889a0;
      --primary:#5c4ee8;
      --primary-soft:#ede9ff;
      --primary-dim:#8b7ff0;
      --accent-r:#e05555;
      --accent-g:#00a878;
      --accent-y:#e8a020;
      --accent-b:#3b82f6;
      --header-bg:#ffffff;
      --header-border:#dde1e7;
      --kpi-bg:#ffffff;
      --tag-bg:#ede9ff;
      --tag-color:#5c4ee8;
      --gantt-bg:#f8f9fb;
      --gantt-head:#eef0f4;
      --today-line:#e05555;
      --modal-overlay:rgba(0,0,0,0.35);
      --scrollbar:#c8ccd8;
      --scrollbar-hover:#a0a5c0;
    }
    [data-theme="dark"] {
      --bg:#0f0e1a;
      --surface:#1a1830;
      --surface2:#14122a;
      --border:#2d2850;
      --text:#e8e4ff;
      --text2:#a29bfe;
      --text3:#636e72;
      --primary:#6c5ce7;
      --primary-soft:#2d2850;
      --primary-dim:#9d8fff;
      --accent-r:#e17055;
      --accent-g:#00b894;
      --accent-y:#fdcb6e;
      --accent-b:#74b9ff;
      --header-bg:#1a1830;
      --header-border:#2d2850;
      --kpi-bg:#14122a;
      --tag-bg:#2d2850;
      --tag-color:#a29bfe;
      --gantt-bg:#14122a;
      --gantt-head:#1a1830;
      --today-line:#e17055;
      --modal-overlay:rgba(0,0,0,0.6);
      --scrollbar:#3d3870;
      --scrollbar-hover:#6c5ce7;
    }

    body{background:var(--bg);color:var(--text);font-family:'Apple SD Gothic Neo','Malgun Gothic',-apple-system,BlinkMacSystemFont,sans-serif;font-size:14px;line-height:1.5;min-height:100vh;transition:background .2s,color .2s;}
    a{color:inherit;text-decoration:none;}
    ::-webkit-scrollbar{width:5px;height:5px;}
    ::-webkit-scrollbar-track{background:var(--surface2);}
    ::-webkit-scrollbar-thumb{background:var(--scrollbar);border-radius:3px;}
    ::-webkit-scrollbar-thumb:hover{background:var(--scrollbar-hover);}

    /* ── HEADER ── */
    .g-header{background:var(--header-bg);border-bottom:1px solid var(--header-border);padding:0 20px;height:52px;display:flex;align-items:center;gap:12px;position:sticky;top:0;z-index:300;box-shadow:0 1px 4px rgba(0,0,0,.06);}
    .logo{font-size:15px;font-weight:900;color:var(--primary);letter-spacing:-.3px;white-space:nowrap;}
    .logo span{color:var(--text);}
    .h-spacer{flex:1;}
    .h-stat{font-size:11px;color:var(--text3);white-space:nowrap;}
    .h-stat b{color:var(--primary);}
    .theme-btn{background:var(--surface2);border:1px solid var(--border);border-radius:20px;padding:4px 12px;font-size:11px;font-weight:700;color:var(--text2);cursor:pointer;transition:all .15s;white-space:nowrap;}
    .theme-btn:hover{background:var(--primary);color:#fff;border-color:var(--primary);}

    /* ── KPI BAR ── */
    .kpi-bar{background:var(--surface);border-bottom:1px solid var(--border);padding:0 20px;display:flex;gap:0;position:sticky;top:52px;z-index:290;overflow-x:auto;}
    .kpi-bar::-webkit-scrollbar{height:0;}
    .kpi-item{display:flex;flex-direction:column;align-items:center;padding:8px 24px;border-right:1px solid var(--border);min-width:110px;cursor:pointer;transition:background .15s;position:relative;}
    .kpi-item:last-child{border-right:none;}
    .kpi-item:hover{background:var(--primary-soft);}
    .kpi-item.active{background:var(--primary-soft);}
    .kpi-label{font-size:10px;font-weight:700;color:var(--text3);white-space:nowrap;margin-bottom:2px;}
    .kpi-val{font-size:24px;font-weight:900;line-height:1.1;}
    .kv-r{color:var(--accent-r);}
    .kv-g{color:var(--accent-g);}
    .kv-y{color:var(--accent-y);}
    .kv-b{color:var(--accent-b);}
    .kpi-panel{display:none;position:absolute;top:100%;left:0;min-width:260px;background:var(--surface);border:1px solid var(--border);border-top:none;border-radius:0 0 8px 8px;box-shadow:0 8px 24px rgba(0,0,0,.12);z-index:400;max-height:320px;overflow-y:auto;}
    .kpi-item.active .kpi-panel{display:block;}
    .kpi-panel-title{font-size:10px;font-weight:700;color:var(--text3);padding:8px 14px 4px;border-bottom:1px solid var(--border);}
    .kpi-panel-item{padding:7px 14px;font-size:12px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:8px;}
    .kpi-panel-item:last-child{border-bottom:none;}
    .kpi-panel-item .game-name{font-weight:600;color:var(--text);flex:1;}
    .kpi-panel-item .game-date{font-size:11px;color:var(--text3);}
    .kpi-panel-item .game-badge{font-size:10px;font-weight:700;padding:1px 6px;border-radius:10px;background:var(--tag-bg);color:var(--tag-color);white-space:nowrap;}
    .kpi-empty{padding:14px;font-size:12px;color:var(--text3);text-align:center;}

    /* ── MONTH NAV BAR ── */
    .month-nav{background:var(--surface2);border-bottom:1px solid var(--border);padding:6px 20px;display:flex;align-items:center;gap:8px;position:sticky;top:102px;z-index:280;}
    .month-nav-btn{background:var(--surface);border:1px solid var(--border);border-radius:5px;padding:4px 10px;font-size:12px;font-weight:700;color:var(--text2);cursor:pointer;transition:all .15s;}
    .month-nav-btn:hover{background:var(--primary);color:#fff;border-color:var(--primary);}
    .month-select{background:var(--surface);border:1px solid var(--border);border-radius:5px;padding:4px 8px;font-size:12px;font-weight:700;color:var(--text);outline:none;cursor:pointer;}
    .month-select:focus{border-color:var(--primary);}
    .month-label{font-size:13px;font-weight:800;color:var(--primary);}
    .mns{flex:1;}

    /* ── MAIN TABS ── */
    .tab-nav{background:var(--surface);border-bottom:1px solid var(--border);padding:0 20px;display:flex;gap:0;position:sticky;top:142px;z-index:270;}
    .tab-btn{padding:10px 20px;font-size:13px;font-weight:700;color:var(--text3);border:none;background:none;cursor:pointer;border-bottom:2px solid transparent;transition:all .15s;}
    .tab-btn:hover{color:var(--text);}
    .tab-btn.active{color:var(--primary);border-bottom-color:var(--primary);}
    .tab-page{display:none;}
    .tab-page.active{display:block;}

    /* ── FILTER BAR ── */
    .filter-bar{background:var(--surface);border-bottom:1px solid var(--border);padding:8px 20px;display:flex;align-items:center;gap:6px;flex-wrap:wrap;}
    .filter-label{font-size:11px;font-weight:700;color:var(--text3);}
    .f-chip{background:var(--surface2);border:1px solid var(--border);border-radius:20px;padding:3px 10px;font-size:11px;font-weight:600;color:var(--text2);cursor:pointer;transition:all .15s;}
    .f-chip:hover{border-color:var(--primary);color:var(--primary);}
    .f-chip.active{background:var(--primary);color:#fff;border-color:var(--primary);}
    .filter-sep{width:1px;height:18px;background:var(--border);margin:0 4px;}

    /* ── GANTT ── */
    .gantt-wrap{overflow-x:auto;margin:12px 20px;border:1px solid var(--border);border-radius:8px;background:var(--gantt-bg);}
    .gantt-head{display:grid;background:var(--gantt-head);border-bottom:1px solid var(--border);height:30px;}
    .gantt-day{font-size:10px;font-weight:700;color:var(--text3);display:flex;align-items:center;justify-content:center;border-right:1px solid var(--border);min-width:0;}
    .gantt-day.today-col{color:var(--accent-r);font-weight:900;}
    .gantt-day:last-child{border-right:none;}
    .gantt-body{position:relative;}
    .gantt-lane{display:grid;height:28px;position:relative;border-bottom:1px solid var(--border);}
    .gantt-lane:last-child{border-bottom:none;}
    .gantt-cell{border-right:1px solid var(--border);opacity:.3;}
    .gantt-cell:last-child{border-right:none;}
    .gantt-cell.today-col{background:var(--accent-r);opacity:.08;}
    .tl-bar{position:absolute;top:4px;height:20px;border-radius:4px;display:flex;align-items:center;overflow:hidden;cursor:pointer;transition:opacity .15s;z-index:10;}
    .tl-bar:hover{opacity:.85;}
    .tl-btxt{font-size:10px;font-weight:700;color:#fff;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;padding:0 6px;text-shadow:0 1px 2px rgba(0,0,0,.3);}
    .tl-today{position:absolute;top:0;bottom:0;width:2px;background:var(--today-line);z-index:20;pointer-events:none;}

    /* ── NEWS LIST ── */
    .news-list{padding:12px 20px;display:flex;flex-direction:column;gap:8px;}
    .news-card{background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:12px 14px;cursor:pointer;transition:all .15s;display:flex;align-items:flex-start;gap:12px;}
    .news-card:hover{border-color:var(--primary);box-shadow:0 2px 8px rgba(92,78,232,.12);}
    .nc-left{flex:1;min-width:0;}
    .nc-tags{display:flex;gap:4px;margin-bottom:5px;flex-wrap:wrap;}
    .nc-tag{font-size:10px;font-weight:700;padding:1px 6px;border-radius:10px;white-space:nowrap;}
    .tag-cat{background:var(--tag-bg);color:var(--tag-color);}
    .tag-genre{background:#e8f4e8;color:#2d7a2d;}
    [data-theme="dark"] .tag-genre{background:#1a3020;color:#4ade80;}
    .tag-src{background:var(--surface2);color:var(--text3);}
    .nc-title{font-size:13px;font-weight:700;color:var(--text);line-height:1.35;margin-bottom:3px;}
    .nc-summary{font-size:12px;color:var(--text2);line-height:1.4;overflow:hidden;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;}
    .nc-date{font-size:11px;color:var(--text3);white-space:nowrap;margin-top:2px;}
    .nc-right{display:flex;flex-direction:column;align-items:flex-end;gap:4px;min-width:60px;}
    .nc-new{font-size:10px;font-weight:900;color:var(--accent-r);background:rgba(224,85,85,.1);padding:1px 6px;border-radius:10px;}

    /* ── INDUSTRY NEWS LIST ── */
    .industry-list{padding:12px 20px;display:flex;flex-direction:column;gap:6px;}
    .ind-card{background:var(--surface);border:1px solid var(--border);border-radius:6px;padding:10px 14px;cursor:pointer;transition:all .15s;}
    .ind-card:hover{border-color:var(--primary);}
    .ind-header{display:flex;align-items:baseline;gap:8px;margin-bottom:3px;}
    .ind-cat{font-size:10px;font-weight:700;color:var(--tag-color);background:var(--tag-bg);padding:1px 6px;border-radius:10px;white-space:nowrap;}
    .ind-title{font-size:13px;font-weight:600;color:var(--text);flex:1;}
    .ind-meta{font-size:11px;color:var(--text3);}

    /* ── MODAL ── */
    .modal-overlay{display:none;position:fixed;inset:0;background:var(--modal-overlay);z-index:900;align-items:center;justify-content:center;}
    .modal-overlay.open{display:flex;}
    .modal{background:var(--surface);border-radius:12px;max-width:680px;width:calc(100% - 32px);max-height:85vh;display:flex;flex-direction:column;overflow:hidden;box-shadow:0 20px 60px rgba(0,0,0,.25);}
    .modal-head{padding:16px 20px;border-bottom:1px solid var(--border);display:flex;align-items:flex-start;gap:12px;}
    .modal-tags{display:flex;gap:4px;flex-wrap:wrap;margin-bottom:6px;}
    .modal-title{font-size:16px;font-weight:800;color:var(--text);line-height:1.3;}
    .modal-close{margin-left:auto;background:none;border:none;font-size:20px;color:var(--text3);cursor:pointer;line-height:1;padding:2px;flex-shrink:0;}
    .modal-close:hover{color:var(--text);}
    .modal-body{padding:16px 20px;overflow-y:auto;flex:1;}
    .modal-meta{font-size:11px;color:var(--text3);margin-bottom:12px;}
    .modal-summary{font-size:13px;color:var(--text2);line-height:1.6;margin-bottom:12px;padding:10px 12px;background:var(--surface2);border-radius:6px;border-left:3px solid var(--primary);}
    .modal-body-text{font-size:13px;color:var(--text);line-height:1.7;white-space:pre-wrap;word-break:break-all;}
    .modal-link{display:inline-block;margin-top:12px;color:var(--primary);font-size:12px;font-weight:600;text-decoration:underline;}

    /* ── SKELETON ── */
    .skeleton{background:linear-gradient(90deg,var(--surface2) 25%,var(--border) 50%,var(--surface2) 75%);background-size:200% 100%;animation:shimmer 1.5s infinite;}
    @keyframes shimmer{0%{background-position:200% 0;}100%{background-position:-200% 0;}}
    .skel-card{border-radius:8px;height:70px;margin-bottom:8px;}

    /* ── EMPTY STATE ── */
    .empty{padding:40px 20px;text-align:center;color:var(--text3);}
    .empty-icon{font-size:36px;margin-bottom:8px;}
    .empty-msg{font-size:13px;}

    /* ── SECTION TITLE ── */
    .sec-title{font-size:12px;font-weight:700;color:var(--text3);padding:10px 20px 4px;text-transform:uppercase;letter-spacing:.5px;}

    /* ── BAR COLORS ── */
    .bc0{background:#5c4ee8;} .bc1{background:#e05555;} .bc2{background:#00a878;}
    .bc3{background:#e8a020;} .bc4{background:#3b82f6;} .bc5{background:#9b59b6;}
    .bc6{background:#e91e63;} .bc7{background:#009688;}
  </style>
</head>
<body>

<!-- ── HEADER ── -->
<header class="g-header">
  <div class="logo">GAME<span>PULSE</span></div>
  <div class="h-spacer"></div>
  <span class="h-stat" id="h-stat">로딩 중...</span>
  <button class="theme-btn" id="theme-btn" onclick="toggleTheme()">🌙 다크</button>
</header>

<!-- ── KPI BAR ── -->
<div class="kpi-bar" id="kpi-bar">
  <div class="kpi-item" id="kpi-launch" onclick="toggleKpi('kpi-launch')">
    <div class="kpi-label">오늘 출시</div>
    <div class="kpi-val kv-r" id="kv-launch">—</div>
    <div class="kpi-panel" id="kp-launch">
      <div class="kpi-panel-title">오늘 출시 게임</div>
      <div id="kpl-launch"><div class="kpi-empty">데이터 로딩 중...</div></div>
    </div>
  </div>
  <div class="kpi-item" id="kpi-cbt" onclick="toggleKpi('kpi-cbt')">
    <div class="kpi-label">진행중 CBT</div>
    <div class="kpi-val kv-g" id="kv-cbt">—</div>
    <div class="kpi-panel" id="kp-cbt">
      <div class="kpi-panel-title">CBT 진행 게임</div>
      <div id="kpl-cbt"><div class="kpi-empty">데이터 로딩 중...</div></div>
    </div>
  </div>
  <div class="kpi-item" id="kpi-week" onclick="toggleKpi('kpi-week')">
    <div class="kpi-label">이번주 신작</div>
    <div class="kpi-val kv-y" id="kv-week">—</div>
    <div class="kpi-panel" id="kp-week">
      <div class="kpi-panel-title">이번주 출시 게임</div>
      <div id="kpl-week"><div class="kpi-empty">데이터 로딩 중...</div></div>
    </div>
  </div>
  <div class="kpi-item" id="kpi-pre" onclick="toggleKpi('kpi-pre')">
    <div class="kpi-label">사전예약중</div>
    <div class="kpi-val kv-b" id="kv-pre">—</div>
    <div class="kpi-panel" id="kp-pre">
      <div class="kpi-panel-title">사전예약 게임</div>
      <div id="kpl-pre"><div class="kpi-empty">데이터 로딩 중...</div></div>
    </div>
  </div>
</div>

<!-- ── MONTH NAV BAR ── -->
<div class="month-nav">
  <button class="month-nav-btn" onclick="shiftMonth(-1)">◀ 이전달</button>
  <select class="month-select" id="month-select" onchange="onMonthSelect()"></select>
  <span class="mns"></span>
  <span class="month-label" id="month-label"></span>
  <button class="month-nav-btn" onclick="shiftMonth(1)">다음달 ▶</button>
</div>

<!-- ── MAIN TABS ── -->
<nav class="tab-nav">
  <button class="tab-btn active" onclick="switchTab('new-release')">🎮 신작 소식</button>
  <button class="tab-btn" onclick="switchTab('industry')">📰 업계 뉴스</button>
</nav>

<!-- ── TAB: 신작 소식 ── -->
<div class="tab-page active" id="tab-new-release">
  <!-- Genre + Event filter -->
  <div class="filter-bar" id="filter-bar">
    <span class="filter-label">장르</span>
    <button class="f-chip active" data-genre="전체" onclick="setGenre(this)">전체</button>
    <button class="f-chip" data-genre="MMORPG" onclick="setGenre(this)">MMORPG</button>
    <button class="f-chip" data-genre="수집형 RPG" onclick="setGenre(this)">수집형RPG</button>
    <button class="f-chip" data-genre="방치형" onclick="setGenre(this)">방치형</button>
    <button class="f-chip" data-genre="전략" onclick="setGenre(this)">전략</button>
    <button class="f-chip" data-genre="슈팅" onclick="setGenre(this)">슈팅</button>
    <button class="f-chip" data-genre="RPG" onclick="setGenre(this)">RPG</button>
    <div class="filter-sep"></div>
    <span class="filter-label">이벤트</span>
    <button class="f-chip active" data-etype="전체" onclick="setEtype(this)">전체</button>
    <button class="f-chip" data-etype="사전예약" onclick="setEtype(this)">사전예약</button>
    <button class="f-chip" data-etype="CBT" onclick="setEtype(this)">CBT</button>
    <button class="f-chip" data-etype="OBT" onclick="setEtype(this)">OBT</button>
    <button class="f-chip" data-etype="출시" onclick="setEtype(this)">출시</button>
    <button class="f-chip" data-etype="신규서버" onclick="setEtype(this)">신규서버</button>
  </div>

  <!-- Gantt -->
  <div class="gantt-wrap" id="gantt-wrap">
    <div class="gantt-head" id="gantt-head"></div>
    <div class="gantt-body" id="gantt-body"></div>
  </div>

  <!-- News list -->
  <div class="sec-title" id="nr-count-label">신작 소식</div>
  <div class="news-list" id="nr-list"></div>
</div>

<!-- ── TAB: 업계 뉴스 ── -->
<div class="tab-page" id="tab-industry">
  <div class="industry-list" id="ind-list"></div>
</div>

<!-- ── MODAL ── -->
<div class="modal-overlay" id="modal" onclick="closeModalOnBg(event)">
  <div class="modal">
    <div class="modal-head">
      <div style="flex:1;min-width:0;">
        <div class="modal-tags" id="modal-tags"></div>
        <div class="modal-title" id="modal-title"></div>
      </div>
      <button class="modal-close" onclick="closeModal()">✕</button>
    </div>
    <div class="modal-body">
      <div class="modal-meta" id="modal-meta"></div>
      <div class="modal-summary" id="modal-summary" style="display:none;"></div>
      <div class="modal-body-text" id="modal-body"></div>
      <a class="modal-link" id="modal-link" href="#" target="_blank" rel="noopener">원문 보기 →</a>
    </div>
  </div>
</div>

<script>
// ── CONFIG ──────────────────────────────────────────────────────────────────
var DATA_PATH = 'data/';
var DATES_URL = DATA_PATH + 'dates.json';

// ── STATE ───────────────────────────────────────────────────────────────────
var ALL_ARTICLES = [];
var DATES_LIST = [];
var CUR_MONTH = { y: new Date().getFullYear(), m: new Date().getMonth() + 1 };
var GENRE_FILTER = '전체';
var ETYPE_FILTER = '전체';
var ACTIVE_TAB = 'new-release';

// ── NEW GAME EVENT TYPES (출시·CBT 계열) ──────────────────────────────────
var NEW_ET = new Set(['사전예약','CBT','OBT','출시','신규서버','얼리액세스','소프트론칭']);
var SAFE_ET = new Set(['사전예약','CBT','OBT','신규서버','얼리액세스']);

// 엄격 제외 키워드
var EXCL_HARD = [
  '신규 캐릭터','신규캐릭터','캐릭터 출시','캐릭터출시',
  '콜라보','콜라버레이션','카페 콜라보','반주년','기념 이벤트','기념일',
  '협업 이벤트','신규 스킨','스킨 출시'
];
var EXCL_SOFT = [
  '업데이트','패치','밸런스','이벤트 시작','이벤트 오픈',
  '시즌 시작','시즌 오픈','점검','신규 콘텐츠','캐릭터 소개','캐릭터 공개'
];

// 업계 뉴스 카테고리 (신작 소식에서 제외)
var IND_CATS = new Set(['업계 동향','IR','인사','M&A','규제','인터뷰','채용','기술']);

// ── THEME ───────────────────────────────────────────────────────────────────
function toggleTheme() {
  var html = document.documentElement;
  var dark = html.getAttribute('data-theme') === 'dark';
  html.setAttribute('data-theme', dark ? 'light' : 'dark');
  document.getElementById('theme-btn').textContent = dark ? '🌙 다크' : '☀️ 라이트';
}

// ── cleanT: prefix 반복 제거 ─────────────────────────────────────────────
function cleanT(t) {
  if (!t) return '';
  var prev;
  do {
    prev = t;
    t = t.replace(/^[\[\(【《〈][^\]\)】》〉]*[\]\)】》〉]\s*/, '');
  } while (t !== prev);
  return t.trim();
}

// ── UTILS ───────────────────────────────────────────────────────────────────
function kst(s) {
  if (!s) return null;
  var d = new Date(s.indexOf('T') > 0 ? s : s + 'T00:00:00+09:00');
  return isNaN(d) ? null : d;
}

function fmtDate(s) {
  if (!s) return '';
  var d = kst(s);
  if (!d) return s;
  return (d.getMonth()+1) + '/' + d.getDate();
}

function today() {
  var n = new Date();
  return new Date(n.getFullYear(), n.getMonth(), n.getDate());
}

function weekRange() {
  var t = today();
  var dow = t.getDay();
  var mon = new Date(t); mon.setDate(t.getDate() - (dow === 0 ? 6 : dow - 1));
  var sun = new Date(mon); sun.setDate(mon.getDate() + 6);
  return { mon: mon, sun: sun };
}

// ── MONTH NAV ───────────────────────────────────────────────────────────────
function buildMonthSelect() {
  var sel = document.getElementById('month-select');
  sel.innerHTML = '';
  for (var y = 2024; y <= 2027; y++) {
    for (var mo = 1; mo <= 12; mo++) {
      var opt = document.createElement('option');
      opt.value = y + '-' + mo;
      opt.textContent = y + '년 ' + mo + '월';
      if (y === CUR_MONTH.y && mo === CUR_MONTH.m) opt.selected = true;
      sel.appendChild(opt);
    }
  }
}

function onMonthSelect() {
  var v = document.getElementById('month-select').value.split('-');
  CUR_MONTH = { y: parseInt(v[0]), m: parseInt(v[1]) };
  updateMonthLabel();
  render();
}

function shiftMonth(d) {
  CUR_MONTH.m += d;
  if (CUR_MONTH.m > 12) { CUR_MONTH.m = 1; CUR_MONTH.y++; }
  if (CUR_MONTH.m < 1)  { CUR_MONTH.m = 12; CUR_MONTH.y--; }
  updateMonthLabel();
  buildMonthSelect();
  render();
}

function updateMonthLabel() {
  document.getElementById('month-label').textContent = CUR_MONTH.y + '년 ' + CUR_MONTH.m + '월';
}

// ── FILTERS ─────────────────────────────────────────────────────────────────
function setGenre(btn) {
  document.querySelectorAll('[data-genre]').forEach(function(b){b.classList.remove('active');});
  btn.classList.add('active');
  GENRE_FILTER = btn.dataset.genre;
  render();
}

function setEtype(btn) {
  document.querySelectorAll('[data-etype]').forEach(function(b){b.classList.remove('active');});
  btn.classList.add('active');
  ETYPE_FILTER = btn.dataset.etype;
  render();
}

// ── EXCLUSION LOGIC ──────────────────────────────────────────────────────────
function isExcluded(a) {
  // 안전한 이벤트 타입은 제외하지 않음
  if (SAFE_ET.has(a.event_type || '')) return false;
  var t = cleanT(a.cleaned_title || a.title || '').toLowerCase();
  // 무조건 제외
  if (EXCL_HARD.some(function(k){ return t.indexOf(k) !== -1; })) return true;
  // event_type 없는 경우에만 소프트 제외
  if (!NEW_ET.has(a.event_type || '') && EXCL_SOFT.some(function(k){ return t.indexOf(k) !== -1; })) return true;
  return false;
}

// ── NEW RELEASE FILTER ───────────────────────────────────────────────────────
function isNewRelease(a) {
  if (IND_CATS.has(a.cat_html || '')) return false;
  if (isExcluded(a)) return false;
  // is_new_event 플래그 OR event_type이 NEW_ET에 속하면 신작
  return a.is_new_event || NEW_ET.has(a.event_type || '');
}

// ── INDUSTRY NEWS FILTER ─────────────────────────────────────────────────────
function isIndustry(a) {
  return !isNewRelease(a);
}

// ── GET ARTICLES FOR CURRENT MONTH ──────────────────────────────────────────
function getMonthArticles() {
  var mStart = new Date(CUR_MONTH.y, CUR_MONTH.m - 1, 1);
  var mEnd = new Date(CUR_MONTH.y, CUR_MONTH.m, 0);

  return ALL_ARTICLES.filter(function(a) {
    // 기사 날짜 기준으로 당월 포함 여부 판단
    var pubD = kst(a.pub_date || a.collected_at || '');
    var esD = kst(a.event_start || '');
    var eeD = kst(a.event_end || '');

    // 이벤트 기간이 당월과 겹치는 경우
    if (esD) {
      var eEnd = eeD || esD;
      if (esD <= mEnd && eEnd >= mStart) return true;
    }
    // 기사 발행일이 당월인 경우
    if (pubD && pubD >= mStart && pubD <= mEnd) return true;
    return false;
  });
}

// ── KPI ─────────────────────────────────────────────────────────────────────
function updateKpi(articles) {
  var t = today();
  var wr = weekRange();

  var launch = [], cbtList = [], weekList = [], preList = [];

  articles.filter(isNewRelease).forEach(function(a) {
    var et = a.event_type || '';
    var esD = kst(a.event_start || '');
    var eeD = kst(a.event_end || '');
    var gn = cleanT(a.cleaned_title || a.title || '');

    if (et === '출시' && esD) {
      var d = new Date(esD.getFullYear(), esD.getMonth(), esD.getDate());
      if (d.getTime() === t.getTime()) launch.push({ name: gn, date: fmtDate(a.event_start), badge: '출시' });
    }
    if ((et === 'CBT' || et === 'OBT') && esD) {
      var es = new Date(esD.getFullYear(), esD.getMonth(), esD.getDate());
      var ee = eeD ? new Date(eeD.getFullYear(), eeD.getMonth(), eeD.getDate()) : es;
      if (es <= t && t <= ee) cbtList.push({ name: gn, date: fmtDate(a.event_start) + (a.event_end ? '~' + fmtDate(a.event_end) : ''), badge: et });
    }
    if ((et === '출시' || et === 'OBT') && esD) {
      var es2 = new Date(esD.getFullYear(), esD.getMonth(), esD.getDate());
      if (es2 >= wr.mon && es2 <= wr.sun) weekList.push({ name: gn, date: fmtDate(a.event_start), badge: et });
    }
    if (et === '사전예약' && esD) {
      var es3 = new Date(esD.getFullYear(), esD.getMonth(), esD.getDate());
      var ee3 = eeD ? new Date(eeD.getFullYear(), eeD.getMonth(), eeD.getDate()) : new Date(9999,0,1);
      if (es3 <= t && t <= ee3) preList.push({ name: gn, date: fmtDate(a.event_start), badge: '사전예약' });
    }
  });

  document.getElementById('kv-launch').textContent = launch.length;
  document.getElementById('kv-cbt').textContent = cbtList.length;
  document.getElementById('kv-week').textContent = weekList.length;
  document.getElementById('kv-pre').textContent = preList.length;

  renderKpiList('kpl-launch', launch);
  renderKpiList('kpl-cbt', cbtList);
  renderKpiList('kpl-week', weekList);
  renderKpiList('kpl-pre', preList);
}

function renderKpiList(id, list) {
  var el = document.getElementById(id);
  if (!list.length) { el.innerHTML = '<div class="kpi-empty">해당 없음</div>'; return; }
  el.innerHTML = list.map(function(g) {
    return '<div class="kpi-panel-item"><span class="game-name">' + esc(g.name) + '</span>' +
      '<span class="game-date">' + esc(g.date) + '</span>' +
      '<span class="game-badge">' + esc(g.badge) + '</span></div>';
  }).join('');
}

function toggleKpi(id) {
  var items = document.querySelectorAll('.kpi-item');
  items.forEach(function(el) {
    if (el.id === id) {
      el.classList.toggle('active');
    } else {
      el.classList.remove('active');
    }
  });
  // 외부 클릭 닫기
  event.stopPropagation();
}

document.addEventListener('click', function() {
  document.querySelectorAll('.kpi-item').forEach(function(el){ el.classList.remove('active'); });
});

// ── GANTT ─────────────────────────────────────────────────────────────────
var BAR_COLORS = ['bc0','bc1','bc2','bc3','bc4','bc5','bc6','bc7'];
var colorMap = {};
var colorIdx = 0;

function gameColor(name) {
  if (!colorMap[name]) {
    colorMap[name] = BAR_COLORS[colorIdx % BAR_COLORS.length];
    colorIdx++;
  }
  return colorMap[name];
}

function assignLanes(events, mStart, mEnd, totalDays) {
  var laneEnds = [];
  events.forEach(function(ev) {
    var esD = kst(ev.event_start);
    var eeD = ev.event_end ? kst(ev.event_end) : esD;
    if (!esD) { ev._s = 1; ev._e = 1; ev.lane = 0; return; }
    var s = esD < mStart ? 1 : esD.getDate();
    var e = eeD > mEnd ? totalDays : eeD.getDate();
    s = Math.max(1, s); e = Math.min(totalDays, e);
    ev._s = s; ev._e = e;
    var placed = false;
    for (var i = 0; i < laneEnds.length; i++) {
      if (laneEnds[i] < s) { ev.lane = i; laneEnds[i] = e; placed = true; break; }
    }
    if (!placed) { ev.lane = laneEnds.length; laneEnds.push(e); }
  });
  return Math.max(1, laneEnds.length);
}

function buildGantt(articles) {
  var mStart = new Date(CUR_MONTH.y, CUR_MONTH.m - 1, 1);
  var mEnd = new Date(CUR_MONTH.y, CUR_MONTH.m, 0);
  var totalDays = mEnd.getDate();
  var todayD = today();
  var todayInMonth = todayD >= mStart && todayD <= mEnd;
  var todayCol = todayInMonth ? todayD.getDate() : -1;

  // 이벤트 기간이 당월과 겹치는 신작만
  var events = articles.filter(function(a) {
    if (!isNewRelease(a)) return false;
    if (!a.event_start) return false;
    var esD = kst(a.event_start);
    var eeD = a.event_end ? kst(a.event_end) : esD;
    if (!esD) return false;
    return esD <= mEnd && eeD >= mStart;
  });

  // 장르/이벤트타입 필터 적용
  events = applyFilters(events);

  var nLanes = assignLanes(events, mStart, mEnd, totalDays);

  // HEAD
  var headEl = document.getElementById('gantt-head');
  headEl.style.gridTemplateColumns = 'repeat(' + totalDays + ', 1fr)';
  headEl.innerHTML = '';
  for (var d = 1; d <= totalDays; d++) {
    var div = document.createElement('div');
    div.className = 'gantt-day' + (d === todayCol ? ' today-col' : '');
    div.textContent = d;
    headEl.appendChild(div);
  }

  // BODY
  var bodyEl = document.getElementById('gantt-body');
  bodyEl.innerHTML = '';
  bodyEl.style.position = 'relative';

  // 레인별 행
  for (var lane = 0; lane < nLanes; lane++) {
    var laneEl = document.createElement('div');
    laneEl.className = 'gantt-lane';
    laneEl.style.gridTemplateColumns = 'repeat(' + totalDays + ', 1fr)';
    for (var d2 = 1; d2 <= totalDays; d2++) {
      var cell = document.createElement('div');
      cell.className = 'gantt-cell' + (d2 === todayCol ? ' today-col' : '');
      laneEl.appendChild(cell);
    }
    bodyEl.appendChild(laneEl);
  }

  // 오늘 세로선
  if (todayCol > 0) {
    var todayLine = document.createElement('div');
    todayLine.className = 'tl-today';
    var pct = ((todayCol - 0.5) / totalDays * 100).toFixed(2);
    todayLine.style.left = pct + '%';
    bodyEl.appendChild(todayLine);
  }

  // 바 렌더
  var laneH = 28;
  events.forEach(function(ev) {
    if (ev._s === undefined) return;
    var leftPct = ((ev._s - 1) / totalDays * 100).toFixed(3);
    var widthPct = ((ev._e - ev._s + 1) / totalDays * 100).toFixed(3);
    var topPx = ev.lane * laneH + 4;
    var gn = cleanT(ev.cleaned_title || ev.title || '');
    var label = (ev.event_type ? '[' + ev.event_type + '] ' : '') + gn;
    var bar = document.createElement('div');
    bar.className = 'tl-bar ' + gameColor(gn);
    bar.style.left = leftPct + '%';
    bar.style.width = 'calc(' + widthPct + '% - 2px)';
    bar.style.top = topPx + 'px';
    bar.innerHTML = '<span class="tl-btxt">' + esc(label) + '</span>';
    (function(article){ bar.addEventListener('click', function(){ openModal(article); }); })(ev);
    bodyEl.appendChild(bar);
  });

  var totalH = nLanes * laneH;
  bodyEl.style.height = totalH + 'px';

  if (!events.length) {
    bodyEl.innerHTML = '<div class="empty" style="height:60px;"><div class="empty-msg">이번 달 해당 이벤트 없음</div></div>';
  }
}

// ── APPLY FILTERS ─────────────────────────────────────────────────────────
function applyFilters(articles) {
  return articles.filter(function(a) {
    if (GENRE_FILTER !== '전체' && (a.genre || '') !== GENRE_FILTER) return false;
    if (ETYPE_FILTER !== '전체' && (a.event_type || '') !== ETYPE_FILTER) return false;
    return true;
  });
}

// ── ARTICLE STORE (for safe onclick) ─────────────────────────────────────────
var _articleStore = [];
function storeArticle(a) {
  _articleStore.push(a);
  return _articleStore.length - 1;
}
function resetStore() { _articleStore = []; }

// ── RENDER ───────────────────────────────────────────────────────────────────
function render() {
  var mArticles = getMonthArticles();

  // KPI: 전체 기사 기준
  updateKpi(ALL_ARTICLES);

  // 신작 소식
  var nrAll = mArticles.filter(isNewRelease);
  var nrFiltered = applyFilters(nrAll);

  resetStore();
  buildGantt(mArticles);
  renderNewsList(nrFiltered);

  // 업계 뉴스
  var indAll = mArticles.filter(isIndustry);
  renderIndustryList(indAll);

  // 헤더 stat
  document.getElementById('h-stat').innerHTML =
    '<b>' + mArticles.length + '</b>건 / ' +
    CUR_MONTH.y + '.' + String(CUR_MONTH.m).padStart(2,'0');

  // count label
  document.getElementById('nr-count-label').textContent =
    '신작 소식 ' + nrFiltered.length + '건';
}

function renderNewsList(articles) {
  var el = document.getElementById('nr-list');
  if (!articles.length) {
    el.innerHTML = '<div class="empty"><div class="empty-icon">🔍</div><div class="empty-msg">해당하는 신작 소식이 없습니다</div></div>';
    return;
  }
  el.innerHTML = articles.map(function(a) {
    var title = cleanT(a.cleaned_title || a.title || '');
    var genre = a.genre || '';
    var et = a.event_type || '';
    var cat = a.cat_html || '';
    var src = a.source_name || '';
    var summary = a.summary || '';
    var dateStr = fmtDate(a.event_start || a.pub_date || a.collected_at || '');
    var isNew = a.is_new_event;
    var idx = storeArticle(a);

    return '<div class="news-card" onclick="openModalByIdx(' + idx + ')">' +
      '<div class="nc-left">' +
        '<div class="nc-tags">' +
          (cat ? '<span class="nc-tag tag-cat">' + esc(cat) + '</span>' : '') +
          (et ? '<span class="nc-tag tag-cat" style="background:var(--primary-soft);color:var(--primary);">' + esc(et) + '</span>' : '') +
          (genre ? '<span class="nc-tag tag-genre">' + esc(genre) + '</span>' : '') +
          (src ? '<span class="nc-tag tag-src">' + esc(src) + '</span>' : '') +
        '</div>' +
        '<div class="nc-title">' + esc(title) + '</div>' +
        (summary ? '<div class="nc-summary">' + esc(summary) + '</div>' : '') +
      '</div>' +
      '<div class="nc-right">' +
        (isNew ? '<span class="nc-new">NEW</span>' : '') +
        '<span class="nc-date">' + esc(dateStr) + '</span>' +
      '</div>' +
    '</div>';
  }).join('');
}

function renderIndustryList(articles) {
  var el = document.getElementById('ind-list');
  if (!articles.length) {
    el.innerHTML = '<div class="empty"><div class="empty-icon">📰</div><div class="empty-msg">이번 달 업계 뉴스가 없습니다</div></div>';
    return;
  }
  el.innerHTML = articles.map(function(a) {
    var title = cleanT(a.cleaned_title || a.title || '');
    var cat = a.cat_html || '';
    var src = a.source_name || '';
    var dateStr = fmtDate(a.pub_date || a.collected_at || '');
    var idx = storeArticle(a);
    return '<div class="ind-card" onclick="openModalByIdx(' + idx + ')">' +
      '<div class="ind-header">' +
        (cat ? '<span class="ind-cat">' + esc(cat) + '</span>' : '') +
        '<span class="ind-title">' + esc(title) + '</span>' +
      '</div>' +
      '<div class="ind-meta">' + esc(src) + (dateStr ? ' · ' + dateStr : '') + '</div>' +
    '</div>';
  }).join('');
}

// ── MODAL ───────────────────────────────────────────────────────────────────
var _curArticle = null;

function openModalByIdx(idx) { openModal(_articleStore[idx]); }

function openModal(a) {
  if (!a) return;
  _curArticle = a;
  var title = cleanT(a.cleaned_title || a.title || '');
  var cat = a.cat_html || '';
  var et = a.event_type || '';
  var genre = a.genre || '';
  var src = a.source_name || '';
  var summary = a.summary || '';
  var body = a.body_text || '';
  var link = a.link || a.url || '';
  var dateStr = '';
  if (a.event_start) dateStr = fmtDate(a.event_start) + (a.event_end ? ' ~ ' + fmtDate(a.event_end) : '');
  else dateStr = fmtDate(a.pub_date || a.collected_at || '');

  document.getElementById('modal-tags').innerHTML =
    (cat ? '<span class="nc-tag tag-cat">' + esc(cat) + '</span>' : '') +
    (et ? '<span class="nc-tag tag-cat" style="background:var(--primary-soft);color:var(--primary);">' + esc(et) + '</span>' : '') +
    (genre ? '<span class="nc-tag tag-genre">' + esc(genre) + '</span>' : '') +
    (src ? '<span class="nc-tag tag-src">' + esc(src) + '</span>' : '');
  document.getElementById('modal-title').textContent = title;
  document.getElementById('modal-meta').textContent = dateStr;

  var summEl = document.getElementById('modal-summary');
  if (summary) { summEl.textContent = summary; summEl.style.display = ''; }
  else { summEl.style.display = 'none'; }

  document.getElementById('modal-body').textContent = body || (summary ? '' : '본문 없음');
  var linkEl = document.getElementById('modal-link');
  if (link) { linkEl.href = link; linkEl.style.display = ''; }
  else { linkEl.style.display = 'none'; }

  document.getElementById('modal').classList.add('open');
  document.body.style.overflow = 'hidden';
}

function closeModal() {
  document.getElementById('modal').classList.remove('open');
  document.body.style.overflow = '';
}

function closeModalOnBg(e) {
  if (e.target === document.getElementById('modal')) closeModal();
}

document.addEventListener('keydown', function(e) {
  if (e.key === 'Escape') closeModal();
});

// ── TAB SWITCH ───────────────────────────────────────────────────────────────
function switchTab(id) {
  ACTIVE_TAB = id;
  document.querySelectorAll('.tab-page').forEach(function(p){ p.classList.remove('active'); });
  document.querySelectorAll('.tab-btn').forEach(function(b){ b.classList.remove('active'); });
  document.getElementById('tab-' + id).classList.add('active');
  event.currentTarget.classList.add('active');
}

// ── ESCAPE ───────────────────────────────────────────────────────────────────
function esc(s) {
  return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

// ── DATA LOAD ────────────────────────────────────────────────────────────────
function showSkeleton() {
  ['nr-list','ind-list'].forEach(function(id) {
    var el = document.getElementById(id);
    el.innerHTML = [1,2,3,4].map(function(){ return '<div class="skeleton skel-card"></div>'; }).join('');
  });
}

async function loadDates() {
  try {
    var r = await fetch(DATES_URL + '?v=' + Date.now());
    if (!r.ok) throw new Error('dates.json not found');
    var data = await r.json();
    DATES_LIST = data.dates || data || [];
    if (!DATES_LIST.length) throw new Error('empty dates');
    return DATES_LIST;
  } catch(e) {
    console.warn('dates.json load failed:', e);
    return [];
  }
}

async function loadAllData() {
  showSkeleton();
  var dates = await loadDates();
  if (!dates.length) {
    document.getElementById('nr-list').innerHTML = '<div class="empty"><div class="empty-icon">⚠️</div><div class="empty-msg">데이터를 불러올 수 없습니다.<br>크롤러를 실행해 주세요.</div></div>';
    return;
  }

  // 최근 90일치 로드 (성능 고려)
  var recent = dates.slice(-90);
  var results = await Promise.allSettled(recent.map(function(dateStr) {
    return fetch(DATA_PATH + dateStr + '.json?v=' + Date.now()).then(function(r){ return r.json(); });
  }));

  var seen = new Set();
  ALL_ARTICLES = [];
  results.forEach(function(r) {
    if (r.status !== 'fulfilled') return;
    var arr = Array.isArray(r.value) ? r.value : (r.value.articles || []);
    arr.forEach(function(a) {
      var key = a.link || a.url || a.title;
      if (key && !seen.has(key)) {
        seen.add(key);
        ALL_ARTICLES.push(a);
      }
    });
  });

  ALL_ARTICLES.sort(function(a, b) {
    var da = new Date(a.pub_date || a.collected_at || 0);
    var db = new Date(b.pub_date || b.collected_at || 0);
    return db - da;
  });

  render();
}

// ── INIT ─────────────────────────────────────────────────────────────────────
buildMonthSelect();
updateMonthLabel();
loadAllData();
</script>
</body>
</html>
"""



def _make_articles_data() -> list:
    """ARTICLES 전역 리스트를 JSON 직렬화용 딕셔너리 리스트로 변환"""
    result = []
    for art in ARTICLES:
        result.append({
            "site":            art.get("site", ""),
            "title":           art.get("title", ""),
            "url":             art.get("url", ""),
            "summary":         art.get("summary", ""),
            "body":            art.get("body_text", ""),
            "category":        art.get("cat_html", "일반"),
            "is_domestic":     art.get("is_domestic", True),
            "is_ruliweb_best": art.get("is_ruliweb_best", False),
            "views":           art.get("views", 0),
            "comments":        art.get("comments", 0),
            "collected_at":    art.get("collected_at", ""),
            "pub_date":        art.get("pub_date", ""),
            "score":           art.get("_score", 0),
            "site_cnt":        art.get("_site_cnt", 1),
            "content_score":   art.get("_content_score", 1),
            "covered_sites":   art.get("_covered_sites", []),
            "event_date":      art.get("event_date", ""),
            "event_type":      art.get("event_type", ""),
            "event_start":     art.get("event_start", ""),
            "event_end":       art.get("event_end", ""),
            "genre":           art.get("genre", ""),
            "cleaned_title":   strip_title_prefix(art.get("title", "")),
            "is_new_event":    art.get("cat_html") == "신작 소식",
        })
    return result


def build_html() -> str:
    """index.html 동적 셸 반환 (데이터는 data/*.json 에서 동적 로드)"""
    return HTML_TEMPLATE


def save_html() -> Path:
    """index.html 을 GITHUB_DIR 에 저장 (날짜별 파일 없음)"""
    GITHUB_DIR.mkdir(parents=True, exist_ok=True)
    index = GITHUB_DIR / "index.html"
    index.write_text(HTML_TEMPLATE, encoding="utf-8")
    print(f"  HTML 저장: {index}")
    return index


def save_json(content_date: datetime) -> Path:
    """수집 기사를 data/YYYY-MM-DD.json 으로 저장"""
    data_dir = GITHUB_DIR / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    date_str  = content_date.strftime("%Y-%m-%d")
    file_path = data_dir / f"{date_str}.json"
    articles  = _make_articles_data()

    file_path.write_text(
        json.dumps(articles, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    print(f"  JSON 저장: {file_path} ({len(articles)}건)")
    return file_path


def update_dates_json(max_days: int = 30):
    """data/dates.json — 저장된 날짜 목록을 최신 순으로 유지 (최대 max_days 일)"""
    data_dir   = GITHUB_DIR / "data"
    dates_file = data_dir / "dates.json"
    data_dir.mkdir(parents=True, exist_ok=True)

    # 기존 JSON 파일 목록 스캔
    existing = sorted(
        [p.stem for p in data_dir.glob("????-??-??.json")],
        reverse=True
    )
    # 최대 max_days 개만 유지
    dates = existing[:max_days]

    # 임시 파일에 먼저 쓴 뒤 교체 → 중단 시 손상 방지
    tmp = dates_file.with_suffix(".tmp")
    tmp.write_text(
        json.dumps({"dates": dates}, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    tmp.replace(dates_file)
    print(f"  dates.json 업데이트: {len(dates)}개 날짜")


_ETYPE_MAP = [
    ("사전예약",   ["사전예약", "사전 예약", "사전등록", "사전 등록"]),
    ("CBT",        ["cbt", "클로즈드 베타", "클베", "비공개 테스트", "비공개베타"]),
    ("OBT",        ["obt", "오픈 베타", "공개 베타"]),
    ("신규서버",   ["신규 서버", "신서버", "새 서버"]),
    ("종료",       ["서비스 종료", "서버 종료", "게임 종료", "서비스종료"]),
    ("얼리액세스", ["얼리 액세스", "얼리엑세스", "early access"]),
    ("출시",       ["정식 출시", "정식출시", "그랜드 오픈", "런칭", "서비스 시작", "정식 서비스"]),
]

def detect_event_type(text: str) -> str:
    t = text.lower()
    for etype, kws in _ETYPE_MAP:
        if any(kw in t for kw in kws):
            return etype
    return "출시"


def extract_event_range(title: str, body: str, ref_year: int):
    """(event_start, event_end) YYYY-MM-DD 반환. 없으면 ('', '')"""
    text = (title or "") + " " + (body or "")[:600]

    def mk(y, mo, d):
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y}-{mo:02d}-{d:02d}"
        return None

    # 명시적 범위: MM월 DD일 ~ MM월 DD일
    m = re.search(
        r"(\d{1,2})\s*월\s*(\d{1,2})\s*일\s*[~\-]\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일",
        text
    )
    if m:
        s = mk(ref_year, int(m.group(1)), int(m.group(2)))
        e = mk(ref_year, int(m.group(3)), int(m.group(4)))
        if s and e and s <= e:
            return s, e

    # 단일 날짜 수집
    dates = []
    for m2 in re.finditer(r"(20\d\d)\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일", text):
        d = mk(int(m2.group(1)), int(m2.group(2)), int(m2.group(3)))
        if d:
            dates.append(d)
    for m2 in re.finditer(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일", text):
        d = mk(ref_year, int(m2.group(1)), int(m2.group(2)))
        if d:
            dates.append(d)
    dates = sorted(set(dates))
    if dates:
        return dates[0], dates[0]
    return "", ""


def extract_event_date(title: str, body: str, ref_year: int) -> str:
    """기사 제목/본문에서 이벤트 발생 날짜 추출 (YYYY-MM-DD 반환, 없으면 "")
    우선순위: 연월일 > 월일 > M/D 슬래시 패턴
    """
    text = (title or "") + " " + (body or "")[:400]
    # 2026년 4월 2일 패턴
    m = re.search(r"(20\d\d)\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일", text)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y}-{mo:02d}-{d:02d}"
    # 4월 2일 패턴 (연도 없음 → ref_year)
    m = re.search(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일", text)
    if m:
        mo, d = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{ref_year}-{mo:02d}-{d:02d}"
    # 4/2 슬래시 패턴
    m = re.search(r"\b(\d{1,2})/(\d{1,2})\b", text)
    if m:
        mo, d = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{ref_year}-{mo:02d}-{d:02d}"
    return ""


def extract_date_from_soup(soup) -> "datetime | None":
    """기사 페이지 HTML에서 발행일 추출 (meta → time → 텍스트 순)"""
    # 1. meta 태그 (가장 신뢰도 높음)
    for attrs in [
        {"property": "article:published_time"},
        {"name": "pubdate"},
        {"name": "article:published"},
        {"property": "og:updated_time"},
        {"name": "Date"},
    ]:
        tag = soup.find("meta", attrs=attrs)
        if tag and tag.get("content"):
            try:
                s = tag["content"].strip()
                dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
                return dt.astimezone(KST)
            except Exception:
                pass
    # 2. <time datetime="...">
    tag = soup.find("time", attrs={"datetime": True})
    if tag:
        try:
            s = tag["datetime"].strip()
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            return dt.astimezone(KST)
        except Exception:
            pass
    # 3. 페이지 상단 텍스트에서 날짜 패턴 (YYYY-MM-DD / YYYY.MM.DD)
    text = soup.get_text()[:4000]
    m = re.search(r"(20\d{2})[.\-](0?[1-9]|1[0-2])[.\-](0?[1-9]|[12]\d|3[01])", text)
    if m:
        try:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return datetime(y, mo, d, 12, 0, tzinfo=KST)
        except Exception:
            pass
    return None


def fetch_article_body(url: str, timeout: int = 10):
    """기사 원문 URL에서 (본문 텍스트, 발행일) 반환"""
    try:
        session = requests.Session()
        session.headers.update(ARTICLE_HEADERS)
        r = session.get(url, timeout=timeout, allow_redirects=True)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "lxml")

        # 발행일 먼저 추출 (태그 제거 전)
        pub_dt = extract_date_from_soup(soup)

        for tag in soup(["script", "style", "nav", "header", "footer", "aside",
                          "figure", "figcaption", ".copyright", ".ad", ".banner"]):
            tag.decompose()
        # 한국 게임 언론사 전용 셀렉터 (우선순위 순)
        SITE_SELECTORS = [
            # 게임조선
            ".article_body", ".articleView", ".view_content",
            # 디스이즈게임
            ".news-content", ".view_article", ".article-text",
            # 게임메카
            ".view_text", ".article-body-content",
            # 인벤
            ".news_content", ".view-content",
            # 게임포커스
            ".article_view", ".news_view", ".news-view",
            # 루리웹
            ".view_content", ".board_main_content",
            # 네이버 뉴스
            ".newsct_article", "#articleBodyContents", ".article-body",
            # 공통
            "article", ".article-body", ".article_body", "#articleBody",
            ".post-content", ".entry-content", ".content-body",
            "[itemprop='articleBody']", ".story-body",
        ]
        for sel in SITE_SELECTORS:
            el = soup.select_one(sel)
            if el:
                text = clean(el.get_text(separator=" "))
                if len(text) > 80:
                    return text[:1200], pub_dt
        # 폴백: 충분히 긴 <p> 태그 모음
        paras = [clean(p.get_text()) for p in soup.select("p")
                 if len(p.get_text().strip()) > 40]
        if paras:
            return " ".join(paras[:6])[:1200], pub_dt
        return "", pub_dt
    except Exception:
        return "", None


def enrich_articles_body(win_start=None, win_end=None):
    """수집된 기사 본문 fetch + 발행일 추출 + 날짜 재필터 (병렬 8 workers)"""
    targets = [a for a in ARTICLES if not a.get("body_text")]
    print(f"  본문 보강 중 ({len(targets)}건, 병렬 8)...")

    def _fetch(art):
        body, pub_dt = fetch_article_body(art["url"])
        if body:
            art["body_text"] = body
        # pub_date 없는 기사(HTML 스크래핑)에 날짜 채우기
        if not art.get("pub_date") and pub_dt:
            art["_pub_datetime"] = pub_dt
            art["pub_date"] = pub_dt.strftime("%Y-%m-%d %H:%M")

    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as ex:
        list(ex.map(_fetch, targets))

    # 날짜 재필터: 본문 fetch 후 날짜 확인된 기사 중 범위 벗어난 것 제거
    if win_start and win_end:
        before = len(ARTICLES)
        ARTICLES[:] = [
            a for a in ARTICLES
            if a.get("_pub_datetime") is None
            or (win_start <= a["_pub_datetime"] <= win_end)
        ]
        removed = before - len(ARTICLES)
        if removed:
            print(f"  날짜 재필터: {removed}건 제거 (범위 외)")

    # ── 이벤트 날짜 + 타입 + 기간 추출 ────────────────────────────────────────
    ref_year = datetime.now(KST).year
    for art in ARTICLES:
        if art.get("_content_score", 0) < 4:
            continue
        t = art.get("title", "")
        b = art.get("body_text", "")
        # 이벤트 타입 (신작 소식 카테고리 기사만)
        if art.get("cat_html") in ("신작 소식", "게임 소식"):
            art["event_type"] = detect_event_type(t + " " + b)
        # 이벤트 날짜(단일) — 기존 로직
        if not art.get("event_date"):
            ev = extract_event_date(t, b, ref_year)
            if ev:
                art["event_date"] = ev
        # 이벤트 기간 (start/end)
        if not art.get("event_start"):
            es, ee = extract_event_range(t, b, ref_year)
            if es:
                art["event_start"] = es
                art["event_end"]   = ee
            elif art.get("event_date"):
                art["event_start"] = art["event_date"]
                art["event_end"]   = art["event_date"]

    filled = sum(1 for a in ARTICLES if a.get("body_text"))
    ev_cnt = sum(1 for a in ARTICLES if a.get("event_date"))
    print(f"  본문 보강 완료: {filled}/{len(ARTICLES)}건 / 이벤트 날짜: {ev_cnt}건")
    # 장르 감지 (본문 보강 후 실행)
    for art in ARTICLES:
        if not art.get("genre"):
            art["genre"] = detect_genre(
                art.get("title", ""),
                art.get("body_text", "") or art.get("summary", "")
            )


def push_github(content_date: datetime):
    try:
        os.chdir(GITHUB_DIR)
        msg = f"Update: {content_date.strftime('%Y-%m-%d')} 뉴스 수집"
        subprocess.run(["git", "add", "."],               check=True, capture_output=True)
        subprocess.run(["git", "commit", "-m", msg],      check=True, capture_output=True)
        subprocess.run(["git", "push", "origin", "main"], check=True, capture_output=True)
        print("  GitHub Pages 배포 완료 ✓")
    except subprocess.CalledProcessError as e:
        print(f"  [WARN] GitHub 푸시 실패: {e.stderr.decode(errors='ignore') if e.stderr else e}")
    except Exception as e:
        print(f"  [WARN] GitHub 배포 오류: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# 메인
# ──────────────────────────────────────────────────────────────────────────────
def main():
    now          = get_now()
    content_date = get_content_date(now)   # 수집 대상일 (전일)
    win_start, win_end = get_time_window(now)

    print(f"=== 게임 업계 동향 크롤러 v3.0 | 실행: {now.strftime('%Y-%m-%d %H:%M')} KST ===")
    print(f"    수집 대상: {win_start.strftime('%Y-%m-%d %H:%M')} ~ {win_end.strftime('%Y-%m-%d %H:%M')} KST")

    # 1. 크롤링 + 분류 (주말 포함 매일 실행)
    run_all(win_start, win_end)

    if not ARTICLES:
        print("[ERROR] 수집된 기사 없음. 종료.")
        return

    # 1-1. 본문 보강 (병렬 fetch)
    print("\n[1-1] 기사 본문 보강")
    enrich_articles_body(win_start, win_end)

    # 2. XLSX 저장 (전일 날짜 기준)
    print("\n[2/4] XLSX 저장")
    save_xlsx(content_date)

    # 3. JSON 저장 + index.html 갱신
    print("\n[3/4] JSON + HTML 저장")
    save_json(content_date)
    update_dates_json()
    save_html()

    # 4. GitHub 배포
    print("\n[4/4] GitHub Pages 배포")
    push_github(content_date)

    print(f"\n=== 완료: {len(ARTICLES)}건 처리 ===")


if __name__ == "__main__":
    main()
